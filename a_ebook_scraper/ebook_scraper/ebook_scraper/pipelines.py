# # Define your item pipelines here
# #
# # Don't forget to add your pipeline to the ITEM_PIPELINES setting
# # See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# # useful for handling different item types with a single interface
# from itemadapter import ItemAdapter

# # pipelines.py
# import openpyxl

# class ExcelPipeline:
#     def open_spider(self, spider):
#         '''
#         open_spider(self, spider):

#         This method is called when the spider starts. We create a new Excel workbook (self.workbook) and set the active sheet (self.sheet).
#         We then write the headers for the Excel sheet (Title, Rating, Price, Stock Status).
#         '''
#         # Create a new workbook and set active sheet
#         self.workbook = openpyxl.Workbook()
#         self.sheet = self.workbook.active
#         self.sheet.title = 'Scraped Ebook Data'
        
#         # Write header row in Excel sheet
#         self.sheet.append(['Title', 'Rating', 'Price', 'Stock Status'])
        
#         # Setting column width
#         self.sheet.column_dimensions['A'].width = 40  # Set width for the 'Title' column
#         # You can also add styles, fonts, and more, as needed, using the openpyxl documentation.


#     def process_item(self, item, spider):
#         '''
#         process_item(self, item, spider):

#         For each scraped item, we add a new row to the Excel sheet using self.sheet.append(). The data is taken from the item dictionary, specifically from the title, rating, price, and stock_status fields.
#         '''
#         # Append item data (title, rating, price, stock_status) to the sheet
#         self.sheet.append([item.get('title'),
#                            item.get('rating'),
#                            item.get('price'),
#                            item.get('stock_status')])
#         return item

#     def close_spider(self, spider):
#         '''
#         close_spider(self, spider):

#         This method is called when the spider finishes. We save the Excel workbook to a file called scraped_data.xlsx.
#         '''
#         # Save the workbook when the spider is closed
#         self.workbook.save('scraped_ebook_data.xlsx')

# pipelines.py
import openpyxl
import os

class ExcelAppendPipeline:
    def open_spider(self, spider):
        # Check if the Excel file exists
        file_exists = os.path.exists('scraped_data.xlsx')

        if file_exists:
            # Load the existing workbook if it exists
            self.workbook = openpyxl.load_workbook('scraped_data.xlsx')
            self.sheet = self.workbook.active
        else:
            # Create a new workbook and set active sheet if it doesn't exist
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = 'Scraped Data'
            # Write header row if file doesn't exist
            self.sheet.append(['Title', 'Rating', 'Price', 'Stock Status'])

    def process_item(self, item, spider):
        # Append item data (title, rating, price, stock_status) to the sheet
        self.sheet.append([item.get('title'),
                           item.get('rating'),
                           item.get('price'),
                           item.get('stock_status')])
        return item

    def close_spider(self, spider):
        # Save the workbook when the spider is closed
        self.workbook.save('scraped_data.xlsx')
